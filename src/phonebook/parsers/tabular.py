"""Parser for tabular Medifox contact exports."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from phonebook.models import (
    ContactRecord,
    CustomerRecord,
    clean,
    compute_customer_key,
    normalize_phone,
    split_name,
)


def find_header_row(ws) -> int | None:
    search_for = {"name", "nachname", "vorname", "telefon", "mobil", "pflegegrad"}
    for row in range(1, min(ws.max_row, 50) + 1):
        row_tokens = {clean(ws.cell(row, column).value).lower() for column in range(1, ws.max_column + 1)}
        if len(search_for.intersection(row_tokens)) >= 2:
            return row
    return None


def parse_tabular(path: Path) -> list[CustomerRecord]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = find_header_row(worksheet)
    if header_row is None:
        raise ValueError("Could not find header row for tabular sheet")

    headers = {clean(worksheet.cell(header_row, column).value).lower(): column for column in range(1, worksheet.max_column + 1)}

    def col(*names: str) -> int | None:
        for name in names:
            if name in headers:
                return headers[name]
        return None

    c_last = col("nachname", "name")
    c_first = col("vorname")
    c_full = col("kunde", "klient")
    c_phone = col("telefon", "tel", "phone")
    c_mobile = col("mobil", "handy", "mobile")
    c_care = col("pflegegrad", "pg")
    c_ext = col("klienten-nr.", "kundennummer", "kunden-nr", "id")
    cc_name = col("kontakt_name", "kontakt", "bezugsperson")
    cc_first = col("kontakt_vorname")
    cc_rel = col("kontakt_relation", "beziehung", "relation")
    cc_phone = col("kontakt_telefon", "kontakt_phone")
    cc_mobile = col("kontakt_mobil", "kontakt_mobile")

    if not any([c_last, c_first, c_full]):
        raise ValueError("No customer name columns found in tabular sheet")

    customers: list[CustomerRecord] = []
    by_key: dict[str, CustomerRecord] = {}
    for row in range(header_row + 1, worksheet.max_row + 1):
        last = clean(worksheet.cell(row, c_last).value) if c_last else ""
        first = clean(worksheet.cell(row, c_first).value) if c_first else ""

        if c_full and not (last or first):
            full_last, full_first = split_name(clean(worksheet.cell(row, c_full).value))
            last = last or full_last
            first = first or full_first

        if not (last or first):
            continue

        record = CustomerRecord(
            lastname=last,
            firstname=first,
            carelevel=clean(worksheet.cell(row, c_care).value) if c_care else "",
            phone=normalize_phone(worksheet.cell(row, c_phone).value) if c_phone else "",
            mobile=normalize_phone(worksheet.cell(row, c_mobile).value) if c_mobile else "",
            external_id=clean(worksheet.cell(row, c_ext).value) if c_ext else "",
        )

        key = compute_customer_key(record)
        if key not in by_key:
            by_key[key] = record
            customers.append(record)

        target = by_key[key]
        contact_last = clean(worksheet.cell(row, cc_name).value) if cc_name else ""
        contact_first = clean(worksheet.cell(row, cc_first).value) if cc_first else ""
        if contact_last or contact_first:
            target.contacts.append(
                ContactRecord(
                    lastname=contact_last,
                    firstname=contact_first,
                    relation=clean(worksheet.cell(row, cc_rel).value) if cc_rel else "",
                    phone=normalize_phone(worksheet.cell(row, cc_phone).value) if cc_phone else "",
                    mobile=normalize_phone(worksheet.cell(row, cc_mobile).value) if cc_mobile else "",
                )
            )

    return customers
