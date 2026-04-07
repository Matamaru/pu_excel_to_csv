"""Parser dispatch for supported Medifox Excel layouts."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from phonebook.models import CustomerRecord, clean
from phonebook.parsers.medifox_report import parse_medifox_report
from phonebook.parsers.tabular import parse_tabular


def detect_sheet_kind(path: Path) -> str:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    sample_text: list[str] = []
    for row in range(1, min(40, worksheet.max_row) + 1):
        for column in range(1, min(25, worksheet.max_column) + 1):
            value = clean(worksheet.cell(row, column).value)
            if value:
                sample_text.append(value)

    blob = " | ".join(sample_text)
    if "Klienten-Nr.:" in blob and "Bezieh.:" in blob:
        return "medifox_report"

    header_tokens = {text.lower() for text in sample_text[:80]}
    if any(token in header_tokens for token in ["name", "nachname"]) and any(
        token in header_tokens for token in ["vorname", "pflegegrad", "telefon", "mobil"]
    ):
        return "tabular_contacts"

    return "unknown"


def parse_excel(path: Path) -> tuple[str, list[CustomerRecord]]:
    kind = detect_sheet_kind(path)
    if kind == "medifox_report":
        return kind, parse_medifox_report(path)
    if kind == "tabular_contacts":
        return kind, parse_tabular(path)
    raise ValueError(f"Unknown sheet layout in {path.name}")


__all__ = ["detect_sheet_kind", "parse_excel", "parse_medifox_report", "parse_tabular"]
