"""Parser for Medifox report-style exports."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from phonebook.models import ContactRecord, CustomerRecord, clean, normalize_phone, split_name


def is_client_start(ws, row: int) -> bool:
    name = clean(ws.cell(row, 2).value)
    marker = clean(ws.cell(row, 16).value)
    return bool(name and "," in name and marker == "Klienten-Nr.:")


def find_next_client_start(ws, row: int) -> int:
    for rr in range(row + 1, ws.max_row + 1):
        if is_client_start(ws, rr):
            return rr
    return ws.max_row + 1


def extract_report_customer(ws, row: int) -> CustomerRecord:
    lastname, firstname = split_name(clean(ws.cell(row, 2).value))
    phone = normalize_phone(ws.cell(row, 9).value)
    external_id = clean(ws.cell(row, 19).value)
    mobile = ""
    carelevel = ""

    for rr in range(row, min(row + 8, ws.max_row + 1)):
        labels = {clean(ws.cell(rr, c).value): c for c in range(1, ws.max_column + 1) if clean(ws.cell(rr, c).value)}
        if "Mobil:" in labels:
            col = labels["Mobil:"]
            if col + 2 <= ws.max_column:
                mobile = mobile or normalize_phone(ws.cell(rr, col + 2).value)
            if not mobile and col + 1 <= ws.max_column:
                mobile = normalize_phone(ws.cell(rr, col + 1).value)
        for key in ("Pflegegrad:", "Pflegegrad", "PG:"):
            if key in labels:
                col = labels[key]
                carelevel = (
                    carelevel
                    or clean(ws.cell(rr, min(col + 3, ws.max_column)).value)
                    or clean(ws.cell(rr, min(col + 1, ws.max_column)).value)
                )

    return CustomerRecord(
        lastname=lastname,
        firstname=firstname,
        carelevel=carelevel,
        phone=phone,
        mobile=mobile,
        external_id=external_id,
    )


def extract_report_contact(ws, row: int) -> ContactRecord:
    lastname, firstname = split_name(clean(ws.cell(row, 3).value))
    relation = clean(ws.cell(row, 7).value)
    phone = normalize_phone(ws.cell(row, 12).value)
    mobile = ""
    for rr in range(row, min(row + 3, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            label = clean(ws.cell(rr, c).value)
            if label == "Mobil:":
                if c + 1 <= ws.max_column:
                    mobile = mobile or normalize_phone(ws.cell(rr, c + 1).value)
                if c + 2 <= ws.max_column:
                    mobile = mobile or normalize_phone(ws.cell(rr, c + 2).value)
    return ContactRecord(lastname=lastname, firstname=firstname, relation=relation, phone=phone, mobile=mobile)


def parse_medifox_report(path: Path) -> list[CustomerRecord]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    customers: list[CustomerRecord] = []
    for start_row in range(1, worksheet.max_row + 1):
        if not is_client_start(worksheet, start_row):
            continue
        customer = extract_report_customer(worksheet, start_row)
        block_end = find_next_client_start(worksheet, start_row) - 1
        for row in range(start_row + 1, block_end + 1):
            if clean(worksheet.cell(row, 6).value) == "Bezieh.:":
                customer.contacts.append(extract_report_contact(worksheet, row))
        customers.append(customer)
    return customers
