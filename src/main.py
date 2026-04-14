#!/usr/bin/env python3
"""Terminal tool to import Medifox Excel exports into SQLite and export phone list CSV."""

from __future__ import annotations

import csv
import hashlib
import os
import shutil
import sqlite3
import sys
from dataclasses import dataclass, field
from pathlib import Path
import re

import openpyxl

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    simpledialog = None


NBSP = "\xa0"
DB_PATH = Path("phonebook.db")
UPLOADS_DIR = Path("uploads")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}


@dataclass
class ContactRecord:
    lastname: str = ""
    firstname: str = ""
    relation: str = ""
    phone: str = ""
    mobile: str = ""


@dataclass
class CustomerRecord:
    lastname: str = ""
    firstname: str = ""
    carelevel: str = ""
    phone: str = ""
    mobile: str = ""
    external_id: str = ""
    contacts: list[ContactRecord] = field(default_factory=list)


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace(NBSP, " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_phone(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    # Keep only useful phone characters for matching while preserving leading +.
    text = text.replace("(0)", "0")
    return re.sub(r"[^0-9+/ -]", "", text).strip()


def normalized_phone_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def split_name(value: str) -> tuple[str, str]:
    value = clean(value)
    if not value:
        return "", ""
    if "," in value:
        last, first = value.split(",", 1)
        return clean(last), clean(first)
    parts = value.split()
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def compute_customer_key(rec: CustomerRecord) -> str:
    raw = "|".join(
        [
            rec.external_id.lower(),
            rec.lastname.lower(),
            rec.firstname.lower(),
            normalized_phone_digits(rec.phone),
            normalized_phone_digits(rec.mobile),
        ]
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_key TEXT NOT NULL UNIQUE,
            external_id TEXT,
            lastname TEXT NOT NULL,
            firstname TEXT NOT NULL,
            carelevel TEXT,
            phone TEXT,
            mobile TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            lastname TEXT NOT NULL,
            firstname TEXT NOT NULL,
            relation TEXT,
            phone TEXT,
            mobile TEXT,
            phone_digits TEXT,
            mobile_digits TEXT,
            FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_customers_phone_digits
            ON customers ((replace(replace(replace(replace(phone, ' ', ''), '-', ''), '/', ''), '+', '')));

        CREATE INDEX IF NOT EXISTS idx_customers_mobile_digits
            ON customers ((replace(replace(replace(replace(mobile, ' ', ''), '-', ''), '/', ''), '+', '')));

        CREATE INDEX IF NOT EXISTS idx_contacts_phone_digits ON contacts(phone_digits);
        CREATE INDEX IF NOT EXISTS idx_contacts_mobile_digits ON contacts(mobile_digits);
        """
    )


def detect_sheet_kind(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sample_text = []
    for r in range(1, min(40, ws.max_row) + 1):
        for c in range(1, min(25, ws.max_column) + 1):
            val = clean(ws.cell(r, c).value)
            if val:
                sample_text.append(val)

    blob = " | ".join(sample_text)
    if "Klienten-Nr.:" in blob and "Bezieh.:" in blob:
        return "medifox_report"

    header_tokens = {s.lower() for s in sample_text[:80]}
    if any(t in header_tokens for t in ["name", "nachname"]) and any(
        t in header_tokens for t in ["vorname", "pflegegrad", "telefon", "mobil"]
    ):
        return "tabular_contacts"

    return "unknown"


def is_client_start(ws, row: int) -> bool:
    name = clean(ws.cell(row, 2).value)
    marker = clean(ws.cell(row, 16).value)
    return bool(name and "," in name and marker == "Klienten-Nr.:")


def find_next_client_start(ws, row: int) -> int:
    for rr in range(row + 1, ws.max_row + 1):
        if is_client_start(ws, rr):
            return rr
    return ws.max_row + 1


def extract_report_customer(ws, row: int) -> CustomerRecord:
    lastname, firstname = split_name(clean(ws.cell(row, 2).value))
    phone = normalize_phone(ws.cell(row, 9).value)
    external_id = clean(ws.cell(row, 19).value)
    mobile = ""
    carelevel = ""

    for rr in range(row, min(row + 8, ws.max_row + 1)):
        labels = {clean(ws.cell(rr, c).value): c for c in range(1, ws.max_column + 1) if clean(ws.cell(rr, c).value)}
        if "Mobil:" in labels:
            col = labels["Mobil:"]
            if col + 2 <= ws.max_column:
                mobile = mobile or normalize_phone(ws.cell(rr, col + 2).value)
            if not mobile and col + 1 <= ws.max_column:
                mobile = normalize_phone(ws.cell(rr, col + 1).value)
        for key in ("Pflegegrad:", "Pflegegrad", "PG:"):
            if key in labels:
                col = labels[key]
                carelevel = (
                    carelevel
                    or clean(ws.cell(rr, min(col + 3, ws.max_column)).value)
                    or clean(ws.cell(rr, min(col + 1, ws.max_column)).value)
                )

    return CustomerRecord(
        lastname=lastname,
        firstname=firstname,
        carelevel=carelevel,
        phone=phone,
        mobile=mobile,
        external_id=external_id,
    )


def extract_report_contact(ws, row: int) -> ContactRecord:
    lastname, firstname = split_name(clean(ws.cell(row, 3).value))
    relation = clean(ws.cell(row, 7).value)
    phone = normalize_phone(ws.cell(row, 12).value)
    mobile = ""
    for rr in range(row, min(row + 3, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            label = clean(ws.cell(rr, c).value)
            if label == "Mobil:":
                if c + 1 <= ws.max_column:
                    mobile = mobile or normalize_phone(ws.cell(rr, c + 1).value)
                if c + 2 <= ws.max_column:
                    mobile = mobile or normalize_phone(ws.cell(rr, c + 2).value)
    return ContactRecord(lastname=lastname, firstname=firstname, relation=relation, phone=phone, mobile=mobile)


def parse_medifox_report(path: Path) -> list[CustomerRecord]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    customers: list[CustomerRecord] = []
    for start_row in range(1, ws.max_row + 1):
        if not is_client_start(ws, start_row):
            continue
        customer = extract_report_customer(ws, start_row)
        block_end = find_next_client_start(ws, start_row) - 1
        for rr in range(start_row + 1, block_end + 1):
            if clean(ws.cell(rr, 6).value) == "Bezieh.:":
                customer.contacts.append(extract_report_contact(ws, rr))
        customers.append(customer)
    return customers


def find_header_row(ws) -> int | None:
    search_for = {"name", "nachname", "vorname", "telefon", "mobil", "pflegegrad"}
    for r in range(1, min(ws.max_row, 50) + 1):
        row_tokens = {clean(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)}
        if len(search_for.intersection(row_tokens)) >= 2:
            return r
    return None


def parse_tabular(path: Path) -> list[CustomerRecord]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError("Could not find header row for tabular sheet")

    headers = {clean(ws.cell(header_row, c).value).lower(): c for c in range(1, ws.max_column + 1)}

    def col(*names: str) -> int | None:
        for n in names:
            if n in headers:
                return headers[n]
        return None

    c_last = col("nachname", "name")
    c_first = col("vorname")
    c_full = col("kunde", "klient")
    c_phone = col("telefon", "tel", "phone")
    c_mobile = col("mobil", "handy", "mobile")
    c_care = col("pflegegrad", "pg")
    c_ext = col("klienten-nr.", "kundennummer", "kunden-nr", "id")
    cc_name = col("kontakt_name", "kontakt", "bezugsperson")
    cc_first = col("kontakt_vorname")
    cc_rel = col("kontakt_relation", "beziehung", "relation")
    cc_phone = col("kontakt_telefon", "kontakt_phone")
    cc_mobile = col("kontakt_mobil", "kontakt_mobile")

    if not any([c_last, c_first, c_full]):
        raise ValueError("No customer name columns found in tabular sheet")

    customers: list[CustomerRecord] = []
    by_key: dict[str, CustomerRecord] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        last = clean(ws.cell(r, c_last).value) if c_last else ""
        first = clean(ws.cell(r, c_first).value) if c_first else ""

        if c_full and not (last or first):
            full_last, full_first = split_name(clean(ws.cell(r, c_full).value))
            last = last or full_last
            first = first or full_first

        if not (last or first):
            continue

        rec = CustomerRecord(
            lastname=last,
            firstname=first,
            carelevel=clean(ws.cell(r, c_care).value) if c_care else "",
            phone=normalize_phone(ws.cell(r, c_phone).value) if c_phone else "",
            mobile=normalize_phone(ws.cell(r, c_mobile).value) if c_mobile else "",
            external_id=clean(ws.cell(r, c_ext).value) if c_ext else "",
        )

        key = compute_customer_key(rec)
        if key not in by_key:
            by_key[key] = rec
            customers.append(rec)

        target = by_key[key]
        contact_last = clean(ws.cell(r, cc_name).value) if cc_name else ""
        contact_first = clean(ws.cell(r, cc_first).value) if cc_first else ""
        if contact_last or contact_first:
            target.contacts.append(
                ContactRecord(
                    lastname=contact_last,
                    firstname=contact_first,
                    relation=clean(ws.cell(r, cc_rel).value) if cc_rel else "",
                    phone=normalize_phone(ws.cell(r, cc_phone).value) if cc_phone else "",
                    mobile=normalize_phone(ws.cell(r, cc_mobile).value) if cc_mobile else "",
                )
            )

    return customers


def parse_excel(path: Path) -> tuple[str, list[CustomerRecord]]:
    kind = detect_sheet_kind(path)
    if kind == "medifox_report":
        return kind, parse_medifox_report(path)
    if kind == "tabular_contacts":
        return kind, parse_tabular(path)
    raise ValueError(f"Unknown sheet layout in {path.name}")


def upsert_customer(conn: sqlite3.Connection, rec: CustomerRecord) -> int:
    customer_key = compute_customer_key(rec)
    conn.execute(
        """
        INSERT INTO customers (customer_key, external_id, lastname, firstname, carelevel, phone, mobile, active)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
        ON CONFLICT(customer_key) DO UPDATE SET
            external_id = excluded.external_id,
            lastname = excluded.lastname,
            firstname = excluded.firstname,
            carelevel = excluded.carelevel,
            phone = excluded.phone,
            mobile = excluded.mobile,
            active = 1,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            customer_key,
            rec.external_id,
            rec.lastname,
            rec.firstname,
            rec.carelevel,
            rec.phone,
            rec.mobile,
        ),
    )
    row = conn.execute("SELECT id FROM customers WHERE customer_key = ?", (customer_key,)).fetchone()
    return int(row[0])


def replace_contacts(conn: sqlite3.Connection, customer_id: int, contacts: list[ContactRecord]) -> None:
    conn.execute("DELETE FROM contacts WHERE customer_id = ?", (customer_id,))
    for c in contacts:
        conn.execute(
            """
            INSERT INTO contacts (customer_id, lastname, firstname, relation, phone, mobile, phone_digits, mobile_digits)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_id,
                c.lastname,
                c.firstname,
                c.relation,
                c.phone,
                c.mobile,
                normalized_phone_digits(c.phone),
                normalized_phone_digits(c.mobile),
            ),
        )


def sync_customers(conn: sqlite3.Connection, imported: list[CustomerRecord]) -> tuple[int, int]:
    if not imported:
        return 0, 0

    seen_ids: set[int] = set()
    with conn:
        for rec in imported:
            customer_id = upsert_customer(conn, rec)
            seen_ids.add(customer_id)
            replace_contacts(conn, customer_id, rec.contacts)

        placeholders = ",".join("?" for _ in seen_ids)
        conn.execute(f"UPDATE customers SET active = 0 WHERE id NOT IN ({placeholders})", tuple(seen_ids))
        conn.execute(f"UPDATE customers SET active = 1 WHERE id IN ({placeholders})", tuple(seen_ids))

    total = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    active = conn.execute("SELECT COUNT(*) FROM customers WHERE active = 1").fetchone()[0]
    return int(total), int(active)


def clear_screen() -> None:
    os.system("cls" if os.name == "nt" else "clear")


def pause() -> None:
    input("\nPress Enter to continue...")


def ensure_uploads_dir() -> None:
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def is_excel_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES


def list_upload_excel_files() -> list[Path]:
    ensure_uploads_dir()
    return sorted(
        [p for p in UPLOADS_DIR.iterdir() if is_excel_file(p)],
        key=lambda p: p.name.lower(),
    )


def parse_paths_input(raw: str) -> list[Path]:
    if not raw.strip():
        return []
    result: list[Path] = []
    for token in [t.strip().strip('"') for t in raw.split(",") if t.strip()]:
        path = Path(token)
        if path.is_dir():
            for child in sorted(path.iterdir(), key=lambda p: p.name.lower()):
                if is_excel_file(child):
                    result.append(child)
            continue
        if is_excel_file(path):
            result.append(path)
    # De-duplicate while keeping order.
    seen: set[Path] = set()
    unique: list[Path] = []
    for path in result:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique.append(path)
    return unique


def copy_files_into_uploads(files: list[Path]) -> list[Path]:
    ensure_uploads_dir()
    copied: list[Path] = []
    for source in files:
        dest = UPLOADS_DIR / source.name
        if source.resolve() != dest.resolve():
            shutil.copy2(source, dest)
        copied.append(dest)
    return copied


def choose_upload_files() -> list[Path]:
    files = list_upload_excel_files()
    if not files:
        print(f"No Excel files in {UPLOADS_DIR}.")
        return []

    print("Select file numbers separated by comma, or 'all'.")
    for idx, file_path in enumerate(files, 1):
        print(f"{idx}) {file_path.name}")

    selection = input("Selection: ").strip().lower()
    if selection == "all":
        return files

    chosen: list[Path] = []
    for token in [t.strip() for t in selection.split(",") if t.strip()]:
        if not token.isdigit():
            continue
        index = int(token)
        if 1 <= index <= len(files):
            chosen.append(files[index - 1])
    return chosen


def run_import(conn: sqlite3.Connection, files: list[Path]) -> None:
    summary = {
        "files_ok": 0,
        "files_failed": 0,
        "customers_parsed": 0,
        "total": 0,
        "active": 0,
        "errors": [],
    }

    if not files:
        print("No valid Excel files selected.")
        return summary

    all_records: list[CustomerRecord] = []

    for path in files:
        if not path.exists() or not path.is_file():
            print(f"Skipping missing file: {path}")
            summary["files_failed"] += 1
            summary["errors"].append(f"Missing file: {path}")
            continue
        try:
            kind, records = parse_excel(path)
            print(f"Detected {kind} for {path.name}: {len(records)} customer(s)")
            all_records.extend(records)
            summary["files_ok"] += 1
            summary["customers_parsed"] += len(records)
        except Exception as exc:
            print(f"Failed to parse {path}: {exc}")
            summary["files_failed"] += 1
            summary["errors"].append(f"{path.name}: {exc}")

    if not all_records:
        print("Nothing imported.")
        return summary

    total, active = sync_customers(conn, all_records)
    print(f"Import done. Total in DB: {total}, active customers: {active}")
    summary["total"] = total
    summary["active"] = active
    return summary


def import_excel_flow(conn: sqlite3.Connection) -> None:
    ensure_uploads_dir()
    print("\n=== Import Excel ===")
    print(f"Uploads folder: {UPLOADS_DIR.resolve()}")
    print("1) Import all Excel files from uploads folder")
    print("2) Select Excel file(s) from uploads folder")
    print("3) Add file(s) to uploads folder and import")
    print("4) Import custom path(s) directly")
    print("0) Back")

    choice = input("Select option: ").strip()
    if choice == "1":
        run_import(conn, list_upload_excel_files())
        return
    if choice == "2":
        run_import(conn, choose_upload_files())
        return
    if choice == "3":
        raw = input("Source file path(s), separated by comma: ").strip()
        source_files = parse_paths_input(raw)
        copied = copy_files_into_uploads(source_files)
        if copied:
            print(f"Copied {len(copied)} file(s) to {UPLOADS_DIR}.")
        run_import(conn, copied)
        return
    if choice == "4":
        raw = input("Excel path(s) or folder path(s), separated by comma: ").strip()
        run_import(conn, parse_paths_input(raw))
        return
    if choice == "0":
        return
    print("Invalid choice.")


def summary_text(summary: dict) -> str:
    lines = [
        f"Files processed: {summary['files_ok']}",
        f"Files failed: {summary['files_failed']}",
        f"Customers parsed: {summary['customers_parsed']}",
    ]
    if summary["total"]:
        lines.append(f"DB customers total: {summary['total']}")
        lines.append(f"DB customers active: {summary['active']}")
    if summary["errors"]:
        lines.append("")
        lines.append("Errors:")
        lines.extend(summary["errors"])
    return "\n".join(lines)


def to_e164_de(phone: str) -> str:
    raw = clean(phone)
    if not raw:
        return ""
    compact = re.sub(r"[^0-9+]", "", raw)
    if compact.startswith("+"):
        return "+" + re.sub(r"\D", "", compact[1:])
    if compact.startswith("00"):
        return "+" + re.sub(r"\D", "", compact[2:])
    digits = re.sub(r"\D", "", compact)
    if not digits:
        return ""
    if digits.startswith("49"):
        return "+" + digits
    if digits.startswith("0"):
        return "+49" + digits[1:]
    return "+" + digits


def hallolena_rows(conn: sqlite3.Connection) -> list[tuple[str, str, str, str]]:
    rows = conn.execute(
        """
        SELECT c.firstname, c.lastname, c.phone, c.mobile, p.firstname, p.lastname, p.phone, p.mobile
        FROM customers c
        LEFT JOIN contacts p ON p.customer_id = c.id
        WHERE c.active = 1
        """
    ).fetchall()

    dedup: set[tuple[str, str, str]] = set()
    out: list[tuple[str, str, str, str]] = []

    def add_entry(firstname: str, lastname: str, phone: str) -> None:
        phone_e164 = to_e164_de(phone)
        if not phone_e164:
            return
        key = (phone_e164, clean(firstname), clean(lastname))
        if key in dedup:
            return
        dedup.add(key)
        out.append((phone_e164, clean(firstname), clean(lastname), ""))

    for c_first, c_last, c_phone, c_mobile, p_first, p_last, p_phone, p_mobile in rows:
        add_entry(c_first or "", c_last or "", c_phone or "")
        add_entry(c_first or "", c_last or "", c_mobile or "")
        add_entry(p_first or "", p_last or "", p_phone or "")
        add_entry(p_first or "", p_last or "", p_mobile or "")

    out.sort(key=lambda r: (r[2].lower(), r[1].lower(), r[0]))
    return out


def export_hallolena_csv_to_path(conn: sqlite3.Connection, out_path: Path) -> int:
    rows = hallolena_rows(conn)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["phone_e164", "first_name", "last_name", "email"])
        writer.writerows(rows)
    return len(rows)


def search_matches(conn: sqlite3.Connection, needle: str) -> list[tuple]:
    return conn.execute(
        """
        SELECT
            c.lastname,
            c.firstname,
            c.phone,
            c.mobile,
            c.active,
            NULL AS contact_lastname,
            NULL AS contact_firstname,
            NULL AS relation,
            NULL AS contact_phone,
            NULL AS contact_mobile,
            'customer' AS source
        FROM customers c
        WHERE replace(replace(replace(replace(c.phone, ' ', ''), '-', ''), '/', ''), '+', '') LIKE '%' || ? || '%'
           OR replace(replace(replace(replace(c.mobile, ' ', ''), '-', ''), '/', ''), '+', '') LIKE '%' || ? || '%'

        UNION ALL

        SELECT
            c.lastname,
            c.firstname,
            c.phone,
            c.mobile,
            c.active,
            p.lastname,
            p.firstname,
            p.relation,
            p.phone,
            p.mobile,
            'contact' AS source
        FROM contacts p
        JOIN customers c ON c.id = p.customer_id
        WHERE p.phone_digits LIKE '%' || ? || '%'
           OR p.mobile_digits LIKE '%' || ? || '%'
        ORDER BY lastname, firstname
        """,
        (needle, needle, needle, needle),
    ).fetchall()


def run_tkinter_ui(conn: sqlite3.Connection) -> None:
    ensure_uploads_dir()
    root = tk.Tk()
    root.title("Medifox Phonebook")
    root.geometry("520x320")

    frame = tk.Frame(root, padx=16, pady=16)
    frame.pack(fill="both", expand=True)

    title = tk.Label(frame, text="Medifox Phonebook", font=("Segoe UI", 16, "bold"))
    title.pack(pady=(0, 12))

    subtitle = tk.Label(frame, text=f"Uploads folder: {UPLOADS_DIR.resolve()}", anchor="w", justify="left")
    subtitle.pack(fill="x", pady=(0, 12))

    def pick_upload_files_dialog(files: list[Path]) -> list[Path]:
        selected: list[Path] = []
        if not files:
            return selected

        dialog = tk.Toplevel(root)
        dialog.title("Select Upload Files")
        dialog.geometry("520x360")
        dialog.transient(root)
        dialog.grab_set()

        info = tk.Label(dialog, text="Select one or more files from uploads:", anchor="w")
        info.pack(fill="x", padx=12, pady=(12, 6))

        listbox = tk.Listbox(dialog, selectmode="extended")
        listbox.pack(fill="both", expand=True, padx=12, pady=6)
        for file_path in files:
            listbox.insert("end", file_path.name)

        actions = tk.Frame(dialog)
        actions.pack(fill="x", padx=12, pady=(4, 12))

        def import_selected() -> None:
            for idx in listbox.curselection():
                selected.append(files[int(idx)])
            dialog.destroy()

        def cancel() -> None:
            dialog.destroy()

        tk.Button(actions, text="Import Selected", command=import_selected).pack(side="left")
        tk.Button(actions, text="Cancel", command=cancel).pack(side="right")

        dialog.wait_window()
        return selected

    def import_from_uploads() -> None:
        summary = run_import(conn, list_upload_excel_files())
        messagebox.showinfo("Import Result", summary_text(summary))

    def import_selected_from_uploads() -> None:
        files = list_upload_excel_files()
        if not files:
            messagebox.showinfo("Uploads", f"No Excel files in {UPLOADS_DIR}.")
            return
        selected = pick_upload_files_dialog(files)
        if not selected:
            return
        summary = run_import(conn, selected)
        messagebox.showinfo("Import Result", summary_text(summary))

    def add_and_import() -> None:
        selected = filedialog.askopenfilenames(
            title="Select Excel file(s)",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm *.xls"), ("All files", "*.*")],
        )
        if not selected:
            return
        files = [Path(p) for p in selected if is_excel_file(Path(p))]
        copied = copy_files_into_uploads(files)
        summary = run_import(conn, copied)
        messagebox.showinfo("Import Result", summary_text(summary))

    def search_phone() -> None:
        raw = simpledialog.askstring("Search", "Phone number to search:", parent=root)
        needle = normalized_phone_digits(raw or "")
        if not needle:
            return
        results = search_matches(conn, needle)
        if not results:
            messagebox.showinfo("Search", "No matches found.")
            return
        lines = []
        for row in results:
            cust_last, cust_first, cust_phone, cust_mobile, active, c_last, c_first, relation, c_phone, c_mobile, source = row
            status = "active" if active else "inactive"
            if source == "customer":
                lines.append(f"[{status}] Customer: {cust_last}, {cust_first} | Tel: {cust_phone} | Mobil: {cust_mobile}")
            else:
                lines.append(
                    f"[{status}] Contact: {c_last}, {c_first} ({relation}) | Tel: {c_phone} | Mobil: {c_mobile} for {cust_last}, {cust_first}"
                )
        messagebox.showinfo("Search Results", "\n".join(lines[:60]))

    def export_csv_ui() -> None:
        output = filedialog.asksaveasfilename(
            title="Save active telephone list CSV",
            defaultextension=".csv",
            initialfile="active_telephone_list.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not output:
            return

        out_path = Path(output)
        rows = conn.execute(
            """
            SELECT
                c.lastname,
                c.firstname,
                c.carelevel,
                c.phone,
                c.mobile,
                p.lastname,
                p.firstname,
                p.relation,
                p.phone,
                p.mobile
            FROM customers c
            LEFT JOIN contacts p ON p.customer_id = c.id
            WHERE c.active = 1
            ORDER BY c.lastname, c.firstname, p.lastname, p.firstname
            """
        ).fetchall()

        header = [
            "name",
            "firstname",
            "pflegegrad",
            "phone",
            "mobile",
            "contact_name",
            "contact_firstname",
            "contact_relation",
            "contact_phone",
            "contact_mobile",
        ]

        with out_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(rows)
        messagebox.showinfo("Export", f"CSV exported:\n{out_path}\n\nRows: {len(rows)}")

    def export_hallolena_ui() -> None:
        output = filedialog.asksaveasfilename(
            title="Save HalloLena single phone list CSV",
            defaultextension=".csv",
            initialfile="hallolena_single_phone_list.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not output:
            return
        out_path = Path(output)
        row_count = export_hallolena_csv_to_path(conn, out_path)
        messagebox.showinfo("Export", f"HalloLena CSV exported:\n{out_path}\n\nRows: {row_count}")

    def stats_ui() -> None:
        total = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        active = conn.execute("SELECT COUNT(*) FROM customers WHERE active = 1").fetchone()[0]
        contacts = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
        messagebox.showinfo("Database Stats", f"Customers: {total} total\nCustomers active: {active}\nContacts: {contacts}")

    buttons = [
        ("Import All From uploads", import_from_uploads),
        ("Select From uploads And Import", import_selected_from_uploads),
        ("Add File(s) And Import", add_and_import),
        ("Search By Phone", search_phone),
        ("Export Active CSV", export_csv_ui),
        ("Export HalloLena Single Phone CSV", export_hallolena_ui),
        ("Show DB Stats", stats_ui),
        ("Exit", root.destroy),
    ]

    for text, cmd in buttons:
        tk.Button(frame, text=text, command=cmd, width=36, pady=6).pack(pady=3)

    root.mainloop()


def export_active_csv(conn: sqlite3.Connection) -> None:
    default_path = Path("active_telephone_list.csv")
    raw = input(f"Output CSV path [{default_path}]: ").strip()
    out_path = Path(raw) if raw else default_path

    rows = conn.execute(
        """
        SELECT
            c.lastname,
            c.firstname,
            c.carelevel,
            c.phone,
            c.mobile,
            p.lastname,
            p.firstname,
            p.relation,
            p.phone,
            p.mobile
        FROM customers c
        LEFT JOIN contacts p ON p.customer_id = c.id
        WHERE c.active = 1
        ORDER BY c.lastname, c.firstname, p.lastname, p.firstname
        """
    ).fetchall()

    header = [
        "name",
        "firstname",
        "pflegegrad",
        "phone",
        "mobile",
        "contact_name",
        "contact_firstname",
        "contact_relation",
        "contact_phone",
        "contact_mobile",
    ]

    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    print(f"CSV exported: {out_path} ({len(rows)} row(s))")


def export_hallolena_csv(conn: sqlite3.Connection) -> None:
    default_path = Path("hallolena_single_phone_list.csv")
    raw = input(f"Output CSV path [{default_path}]: ").strip()
    out_path = Path(raw) if raw else default_path
    row_count = export_hallolena_csv_to_path(conn, out_path)
    print(f"HalloLena CSV exported: {out_path} ({row_count} row(s))")


def search_by_phone(conn: sqlite3.Connection) -> None:
    needle = normalized_phone_digits(input("Phone number to search: "))
    if not needle:
        print("No phone number entered.")
        return

    results = search_matches(conn, needle)

    if not results:
        print("No matches found.")
        return

    for row in results:
        cust_last, cust_first, cust_phone, cust_mobile, active, c_last, c_first, relation, c_phone, c_mobile, source = row
        status = "active" if active else "inactive"
        if source == "customer":
            print(f"[{status}] Customer: {cust_last}, {cust_first} | Tel: {cust_phone} | Mobil: {cust_mobile}")
        else:
            print(
                f"[{status}] Contact: {c_last}, {c_first} ({relation}) | Tel: {c_phone} | Mobil: {c_mobile} "
                f"for {cust_last}, {cust_first}"
            )


def show_stats(conn: sqlite3.Connection) -> None:
    total = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    active = conn.execute("SELECT COUNT(*) FROM customers WHERE active = 1").fetchone()[0]
    contacts = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    print(f"Customers: {total} total, {active} active | Contacts: {contacts}")


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    ensure_uploads_dir()

    if tk is not None:
        try:
            run_tkinter_ui(conn)
            conn.close()
            return
        except Exception as exc:
            print(f"GUI startup failed, falling back to terminal mode: {exc}", file=sys.stderr)

    while True:
        clear_screen()
        print("\n=== Medifox Phonebook Menu ===")
        print("1) Upload / Import Excel sheet(s)")
        print("2) Search by phone number")
        print("3) Export active telephone list CSV")
        print("4) Show DB stats")
        print("5) Export HalloLena single phone list CSV")
        print("0) Exit")

        choice = input("Select option: ").strip()
        if choice == "1":
            import_excel_flow(conn)
            pause()
        elif choice == "2":
            search_by_phone(conn)
            pause()
        elif choice == "3":
            export_active_csv(conn)
            pause()
        elif choice == "4":
            show_stats(conn)
            pause()
        elif choice == "5":
            export_hallolena_csv(conn)
            pause()
        elif choice == "0":
            print("Bye.")
            break
        else:
            print("Invalid choice.")
            pause()

    conn.close()


if __name__ == "__main__":
    main()
